#!/usr/bin/env python3
import argparse
import os
import re
import sys
import warnings
from dataclasses import dataclass, field
from datetime import date, datetime
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any, Iterable, Optional

import psycopg2
from openpyxl import load_workbook
from psycopg2.extras import Json


warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")


DEFAULT_WORKBOOK = Path(
    "/home/orl/Shl/КА/MS BI/bsc_model/agent_pack/Реестр_стратегических_документов_Общества.xlsx"
)
DEFAULT_SOURCE_SYSTEM = "STRATEGY_DOC_REGISTRY_XLSX"

BASE_SHEET = "Реестр (база)"
EXCLUDED_SHEET = "Реестр (искл.)"
DELETED_SHEET = "Удаленные док"

BASE_HEADER_ROW = 12
BASE_DATA_ROW = 13
EXCLUDED_HEADER_ROW = 2
EXCLUDED_DATA_ROW = 3
DELETED_HEADER_ROW = 1
DELETED_DATA_ROW = 2

BASE_OBJECT_COLUMNS = range(14, 25)
BASE_STAGE_COLUMNS = range(25, 32)

OUTCOME_MATCHED = "Сопоставлено"
OUTCOME_MATCHED_REVIEW = "Сопоставлено, нужна проверка"
OUTCOME_REVIEW = "Нужна проверка"
OUTCOME_REJECTED = "Отклонено"
OUTCOME_UNMATCHED = "Не сопоставлено"

MATCH_STATUS_AUTO = "Сопоставлено автоматически"
MATCH_STATUS_MANUAL = "Сопоставлено вручную"
MATCH_STATUS_AMBIGUOUS = "Неоднозначно"
MATCH_STATUS_UNMATCHED = "Не сопоставлено"
MATCH_STATUS_REJECTED = "Отклонено"

MATCH_ROLE_PRIMARY = "Основной"
MATCH_ROLE_COMPONENT = "Компонент"
MATCH_ROLE_CANDIDATE = "Кандидат"
MATCH_ROLE_NONE = "Нет сопоставления"

RECORD_SOURCE_AUTO = "Авто"

STOPWORDS = {
    "стратегия",
    "стратегический",
    "стратегическая",
    "программа",
    "план",
    "комплексный",
    "комплексная",
    "развития",
    "развитие",
    "документ",
    "документа",
    "документов",
    "проекта",
    "проект",
    "проектов",
    "общества",
    "период",
    "на",
    "по",
    "для",
    "и",
    "в",
    "с",
    "г",
    "гг",
}

TOKEN_EXPANSIONS = {
    "тд": ["технический", "директорат"],
    "прд": ["производственный", "директорат"],
    "комд": ["коммерческий", "директорат"],
    "кд": ["кадровый", "директорат"],
    "сги": ["служба", "главного", "инженера"],
    "сит": ["служба", "информационных", "технологий"],
    "it": ["информационных", "технологий"],
    "фд": ["финансовый", "директорат"],
    "дпид": ["проектно", "инвестиционной", "деятельности"],
    "дмтс": ["материально", "технического", "снабжения"],
    "дло": ["логистических", "операций"],
    "уко": ["корпоративных", "отношений"],
    "сурин": ["устойчивого", "развития", "и", "новаций"],
    "дэири": ["эксплуатации", "и", "развития", "инфраструктуры"],
    "дкз": ["корпоративной", "защиты"],
    "дпо": ["правового", "обеспечения"],
    "дтср": ["дтср"],
    "тпр": ["тпр"],
    "укпо": ["укпо"],
    "дпто": ["дпто"],
    "нкс": ["нкс"],
    "обтк": ["обтк"],
    "уоуг": ["уоуг"],
    "спг": ["спг"],
    "тон": ["тон"],
    "впу": ["впу"],
}

TOKEN_CANONICAL_MAP = {
    "кадрового": "кадровый",
    "кадровая": "кадровый",
    "кадровый": "кадровый",
    "технического": "технический",
    "техническая": "технический",
    "технический": "технический",
    "коммерческого": "коммерческий",
    "коммерческая": "коммерческий",
    "коммерческий": "коммерческий",
    "финансового": "финансовый",
    "финансовая": "финансовый",
    "финансовый": "финансовый",
    "производственного": "производственный",
    "производственная": "производственный",
    "производственный": "производственный",
    "директората": "директорат",
    "директорат": "директорат",
    "директорате": "директорат",
    "платформы": "платформа",
    "платформе": "платформа",
    "морской": "морской",
    "морские": "морской",
    "объекты": "объект",
    "объектами": "объект",
    "объекта": "объект",
    "объектов": "объект",
    "танкерами": "танкер",
    "танкеров": "танкер",
    "танкеры": "танкер",
    "нефтеналивными": "нефтеналивный",
    "газовозами": "газовоз",
    "газовозы": "газовоз",
    "поставок": "поставка",
    "маркетинга": "маркетинг",
    "сбыта": "сбыт",
    "книги": "книга",
    "книга": "книга",
    "табличные": "табличный",
    "графические": "графический",
    "изменения": "изменение",
    "уточненные": "уточненный",
    "уточнённые": "уточненный",
}

PHRASE_REPLACEMENTS = {
    "пильтун-а": "па а",
    "пильтун-а ": "па а ",
    "пильтун а": "па а",
    "пильтун-б": "па б",
    "пильтун б": "па б",
    "лун-а": "лун а",
    "нкс-2": "нкс 2",
    "спг/тон/впу": "спг тон впу",
    "spg/ton/vpu": "spg ton vpu",
}

LOW_SIGNAL_DOC_PATTERNS = {
    "список исполнителей": 0.28,
    "титул": 0.26,
    "титулы": 0.26,
    "таблица": 0.22,
    "табличный": 0.20,
    "карта": 0.20,
    "графический": 0.18,
    "приложение": 0.16,
    "ведомость": 0.16,
    "папка": 0.14,
    "backup": 0.22,
    "не применимо": 0.24,
}

DEPARTMENT_ALIASES_RAW = {
    "Технический директорат": ("технический", "тд", "технический директорат"),
    "Производственный директорат": ("производственный", "прд", "производственный директорат"),
    "Коммерческий директорат": ("коммерческий", "комд", "коммерческий директорат"),
    "Директорат проектно-инвестиционной деятельности": (
        "дпид",
        "проектно инвестиционной деятельности",
        "директорат проектно инвестиционной деятельности",
    ),
    "Служба Главного инженера": ("сги", "служба главного инженера"),
    "Служба информационных технологий": (
        "сит",
        "it",
        "информационных технологий",
        "служба информационных технологий",
    ),
    "Финансовый директорат": ("фд", "финансовый", "финансовый директорат"),
    "Служба устойчивого развития и новаций": (
        "сурин",
        "устойчивого развития и новаций",
        "служба устойчивого развития и новаций",
    ),
    "Департамент материально-технического снабжения и организации подрядных работ": (
        "дмтс",
        "материально техническое снабжение и организации подрядных работ",
        "департамент материально технического снабжения и организации подрядных работ",
    ),
    "Кадровый директорат": ("кд", "кадровый", "кадровый директорат"),
    "Управление корпоративных отношений": ("уко", "корпоративных отношений", "управление корпоративных отношений"),
    "Департамент логистических операций": ("дло", "логистические операции", "департамент логистических операций"),
    "Департамент эксплуатации и развития инфраструктуры": (
        "дэири",
        "эксплуатации и развития инфраструктуры",
        "департамент эксплуатации и развития инфраструктуры",
    ),
    "Департамент корпоративной защиты": ("дкз", "корпоративной защиты", "департамент корпоративной защиты"),
    "Директорат правового обеспечения": ("дпо", "правового обеспечения", "директорат правового обеспечения"),
}


DEPARTMENT_ALIAS_INDEX = {}
for canonical_name, aliases in DEPARTMENT_ALIASES_RAW.items():
    for alias in aliases:
        DEPARTMENT_ALIAS_INDEX[alias] = canonical_name


@dataclass(frozen=True)
class IntakeDocument:
    doc_id: str
    short_name: str
    full_name: Optional[str]
    department_code: Optional[str]
    horizon_code: Optional[str]
    status_code: Optional[str]
    short_match: str
    full_match: str
    combined_match: str
    token_set: frozenset[str]
    keyword_set: frozenset[str]


@dataclass
class RegistryRow:
    source_sheet: str
    source_row_num: int
    registry_scope: str
    core_registry_flag: bool
    registry_status_raw: Optional[str]
    registry_record_number: Optional[str]
    registry_document_code: Optional[str]
    registry_document_name: str
    registry_strategy_tactic: Optional[str]
    registry_ksu_flag: Optional[str]
    registry_business_process_id: Optional[str]
    registry_business_process_name: Optional[str]
    registry_document_type: Optional[str]
    registry_department_type: Optional[str]
    registry_department_name: Optional[str]
    registry_department_short_name: Optional[str]
    registry_content_summary: Optional[str]
    registry_object_list: Optional[str]
    registry_stage_list: Optional[str]
    registry_document_about: Optional[str]
    registry_goal_text: Optional[str]
    registry_task_text: Optional[str]
    registry_project_link_text: Optional[str]
    registry_project_id_text: Optional[str]
    registry_project_link_alt_text: Optional[str]
    registry_project_id_alt_text: Optional[str]
    registry_incoming_documents: Optional[str]
    registry_impacted_documents: Optional[str]
    registry_input_data_text: Optional[str]
    registry_output_kpi_text: Optional[str]
    registry_trigger_documents: Optional[str]
    registry_information_recipients: Optional[str]
    registry_horizon_level_3: Optional[str]
    registry_horizon_strict: Optional[str]
    registry_period_label: Optional[str]
    registry_validity_text: Optional[str]
    registry_update_frequency_text: Optional[str]
    registry_actualization_date_raw: Optional[str]
    registry_status_text: Optional[str]
    registry_approved_by_position: Optional[str]
    registry_approved_by_name: Optional[str]
    registry_contact_person: Optional[str]
    registry_exclusion_reason: Optional[str]
    registry_update_plan_text: Optional[str]
    row_payload: dict[str, Any] = field(default_factory=dict)
    canonical_department: Optional[str] = None
    match_name_norm: str = ""
    token_set: frozenset[str] = field(default_factory=frozenset)
    keyword_set: frozenset[str] = field(default_factory=frozenset)


@dataclass
class XwalkRow:
    registry_raw_id: Optional[str]
    source_system: str
    source_snapshot_label: Optional[str]
    technical_snapshot_date: date
    run_id: str
    registry_scope: str
    match_status: str
    match_method: str
    match_role: str
    candidate_rank: Optional[int]
    match_score: Optional[float]
    review_required_flag: bool
    doc_id: Optional[str]
    intake_document_short_name: Optional[str]
    intake_document_full_name: Optional[str]
    intake_department_code: Optional[str]
    intake_horizon_code: Optional[str]
    intake_status_code: Optional[str]
    match_comment: Optional[str]
    record_source: str = RECORD_SOURCE_AUTO


def usage() -> None:
    print(
        "Usage:\n"
        "  python3 ./scripts/strategy-registry-overlay-load.py dry-run [--workbook path] [--run-id run]\n"
        "  python3 ./scripts/strategy-registry-overlay-load.py load [--workbook path] [--run-id run]"
    )


def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(add_help=False)
    parser.add_argument("command", nargs="?", choices=("dry-run", "load"))
    parser.add_argument("--workbook", default=str(DEFAULT_WORKBOOK))
    parser.add_argument("--source-system", default=DEFAULT_SOURCE_SYSTEM)
    parser.add_argument("--source-snapshot-label")
    parser.add_argument("--technical-snapshot-date")
    parser.add_argument("--run-id")
    parser.add_argument("--help", action="store_true")
    args = parser.parse_args(argv[1:])
    if args.help or not args.command:
        usage()
        raise SystemExit(0)
    return args


def require_database_url() -> str:
    url = (os.environ.get("DATABASE_URL") or "").strip()
    if not url:
        raise RuntimeError("DATABASE_URL is required")
    return url


def normalize_text(value: Any) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).replace("\xa0", " ").replace("\r\n", "\n").replace("\r", "\n")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    text = text.strip()
    return text or None


def jsonable_value(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, (bool, int, float)) or value is None:
        return value
    return normalize_text(value)


def normalize_match_text(value: Any) -> str:
    text = normalize_text(value) or ""
    text = text.lower().replace("ё", "е")
    for source, target in PHRASE_REPLACEMENTS.items():
        text = text.replace(source, target)
    text = re.sub(r"[^0-9a-zа-я]+", " ", text)

    tokens: list[str] = []
    for raw_token in text.split():
        expanded = TOKEN_EXPANSIONS.get(raw_token)
        if expanded:
            candidate_tokens = expanded
        else:
            candidate_tokens = [TOKEN_CANONICAL_MAP.get(raw_token, raw_token)]
        for token in candidate_tokens:
            canonical_token = TOKEN_CANONICAL_MAP.get(token, token)
            if canonical_token:
                tokens.append(canonical_token)

    return " ".join(tokens)


def token_set(value: Any) -> frozenset[str]:
    return frozenset(token for token in normalize_match_text(value).split() if len(token) >= 2)


def keyword_set(value: Any) -> frozenset[str]:
    return frozenset(
        token for token in token_set(value) if len(token) >= 3 and token not in STOPWORDS
    )


def canonical_horizon(value: Any) -> Optional[str]:
    text = normalize_match_text(value)
    if not text:
        return None
    if "долгоср" in text:
        return "Долгосрочный"
    if "среднеср" in text:
        return "Среднесрочный"
    if "краткоср" in text:
        return "Краткосрочный"
    if "операц" in text:
        return "Операционный"
    return None


def canonical_department(*raw_values: Any) -> Optional[str]:
    normalized_values = []
    for raw_value in raw_values:
        normalized = normalize_match_text(raw_value)
        if normalized:
            normalized_values.append(normalized)
            if normalized in DEPARTMENT_ALIAS_INDEX:
                return DEPARTMENT_ALIAS_INDEX[normalized]

    for normalized in normalized_values:
        padded = f" {normalized} "
        for alias, canonical in DEPARTMENT_ALIAS_INDEX.items():
            if f" {alias} " in padded:
                return canonical
    return None


def truthy_cell(value: Any) -> bool:
    text = normalize_text(value)
    if text is None:
        return False
    return text not in {"-", "0", "нет", "Нет"}


def make_payload(headers: dict[int, Optional[str]], values: dict[int, Any]) -> dict[str, Any]:
    payload: dict[str, Any] = {}
    for col_idx, raw_value in values.items():
        header = normalize_text(headers.get(col_idx)) or f"column_{col_idx}"
        payload[header] = jsonable_value(raw_value)
    return payload


def load_workbook_rows(workbook_path: Path) -> list[RegistryRow]:
    workbook = load_workbook(workbook_path, data_only=True)
    registry_rows: list[RegistryRow] = []

    registry_rows.extend(load_base_rows(workbook[BASE_SHEET]))
    registry_rows.extend(load_excluded_rows(workbook[EXCLUDED_SHEET]))
    registry_rows.extend(load_deleted_rows(workbook[DELETED_SHEET]))

    return registry_rows


def load_base_rows(sheet) -> list[RegistryRow]:
    headers = {col_idx: sheet.cell(BASE_HEADER_ROW, col_idx).value for col_idx in range(1, sheet.max_column + 1)}
    rows: list[RegistryRow] = []

    for row_num in range(BASE_DATA_ROW, sheet.max_row + 1):
        values = {col_idx: sheet.cell(row_num, col_idx).value for col_idx in range(1, sheet.max_column + 1)}
        document_name = normalize_text(values.get(4))
        if not document_name:
            continue

        object_list = "; ".join(
            normalize_text(headers[col_idx]) or f"column_{col_idx}"
            for col_idx in BASE_OBJECT_COLUMNS
            if truthy_cell(values.get(col_idx))
        ) or None
        stage_list = "; ".join(
            normalize_text(headers[col_idx]) or f"column_{col_idx}"
            for col_idx in BASE_STAGE_COLUMNS
            if truthy_cell(values.get(col_idx))
        ) or None
        payload = make_payload(headers, values)
        payload["object_columns"] = [normalize_text(headers[col_idx]) for col_idx in BASE_OBJECT_COLUMNS if truthy_cell(values.get(col_idx))]
        payload["stage_columns"] = [normalize_text(headers[col_idx]) for col_idx in BASE_STAGE_COLUMNS if truthy_cell(values.get(col_idx))]

        row = RegistryRow(
            source_sheet=BASE_SHEET,
            source_row_num=row_num,
            registry_scope="База",
            core_registry_flag=True,
            registry_status_raw=normalize_text(values.get(1)),
            registry_record_number=normalize_text(values.get(2)),
            registry_document_code=normalize_text(values.get(3)),
            registry_document_name=document_name,
            registry_strategy_tactic=normalize_text(values.get(5)),
            registry_ksu_flag=normalize_text(values.get(6)),
            registry_business_process_id=normalize_text(values.get(7)),
            registry_business_process_name=normalize_text(values.get(8)),
            registry_document_type=normalize_text(values.get(9)),
            registry_department_type=normalize_text(values.get(10)),
            registry_department_name=normalize_text(values.get(11)),
            registry_department_short_name=normalize_text(values.get(12)),
            registry_content_summary=normalize_text(values.get(13)),
            registry_object_list=object_list,
            registry_stage_list=stage_list,
            registry_document_about=normalize_text(values.get(32)),
            registry_goal_text=normalize_text(values.get(33)),
            registry_task_text=normalize_text(values.get(34)),
            registry_project_link_text=normalize_text(values.get(35)),
            registry_project_id_text=normalize_text(values.get(36)),
            registry_project_link_alt_text=normalize_text(values.get(37)),
            registry_project_id_alt_text=normalize_text(values.get(38)),
            registry_incoming_documents=normalize_text(values.get(39)),
            registry_impacted_documents=normalize_text(values.get(40)),
            registry_input_data_text=normalize_text(values.get(41)),
            registry_output_kpi_text=normalize_text(values.get(42)),
            registry_trigger_documents=normalize_text(values.get(43)),
            registry_information_recipients=normalize_text(values.get(44)),
            registry_horizon_level_3=normalize_text(values.get(45)),
            registry_horizon_strict=normalize_text(values.get(46)),
            registry_period_label=normalize_text(values.get(47)),
            registry_validity_text=normalize_text(values.get(48)),
            registry_update_frequency_text=normalize_text(values.get(49)),
            registry_actualization_date_raw=normalize_text(values.get(50)),
            registry_status_text=normalize_text(values.get(51)),
            registry_approved_by_position=normalize_text(values.get(52)),
            registry_approved_by_name=normalize_text(values.get(53)),
            registry_contact_person=normalize_text(values.get(54)),
            registry_exclusion_reason=None,
            registry_update_plan_text=None,
            row_payload=payload,
        )
        enrich_registry_row(row)
        rows.append(row)

    return rows


def load_excluded_rows(sheet) -> list[RegistryRow]:
    headers = {col_idx: sheet.cell(EXCLUDED_HEADER_ROW, col_idx).value for col_idx in range(1, sheet.max_column + 1)}
    rows: list[RegistryRow] = []

    for row_num in range(EXCLUDED_DATA_ROW, sheet.max_row + 1):
        values = {col_idx: sheet.cell(row_num, col_idx).value for col_idx in range(1, sheet.max_column + 1)}
        document_name = normalize_text(values.get(3))
        if not document_name:
            continue

        payload = make_payload(headers, values)
        row = RegistryRow(
            source_sheet=EXCLUDED_SHEET,
            source_row_num=row_num,
            registry_scope="Исключено",
            core_registry_flag=False,
            registry_status_raw=normalize_text(values.get(1)),
            registry_record_number=None,
            registry_document_code=normalize_text(values.get(2)),
            registry_document_name=document_name,
            registry_strategy_tactic=normalize_text(values.get(5)),
            registry_ksu_flag=None,
            registry_business_process_id=normalize_text(values.get(32)),
            registry_business_process_name=normalize_text(values.get(33)),
            registry_document_type=normalize_text(values.get(6)),
            registry_department_type=normalize_text(values.get(7)),
            registry_department_name=normalize_text(values.get(8)),
            registry_department_short_name=None,
            registry_content_summary=normalize_text(values.get(9)),
            registry_object_list=normalize_text(values.get(10)),
            registry_stage_list=normalize_text(values.get(11)),
            registry_document_about=normalize_text(values.get(12)),
            registry_goal_text=normalize_text(values.get(13)),
            registry_task_text=normalize_text(values.get(14)),
            registry_project_link_text=normalize_text(values.get(15)),
            registry_project_id_text=normalize_text(values.get(16)),
            registry_project_link_alt_text=None,
            registry_project_id_alt_text=None,
            registry_incoming_documents=normalize_text(values.get(17)),
            registry_impacted_documents=normalize_text(values.get(19)),
            registry_input_data_text=normalize_text(values.get(21)),
            registry_output_kpi_text=normalize_text(values.get(22)),
            registry_trigger_documents=normalize_text(values.get(23)),
            registry_information_recipients=None,
            registry_horizon_level_3=normalize_text(values.get(24)),
            registry_horizon_strict=normalize_text(values.get(25)),
            registry_period_label=normalize_text(values.get(26)),
            registry_validity_text=normalize_text(values.get(27)),
            registry_update_frequency_text=normalize_text(values.get(28)),
            registry_actualization_date_raw=normalize_text(values.get(29)),
            registry_status_text=normalize_text(values.get(31)),
            registry_approved_by_position=normalize_text(values.get(35)),
            registry_approved_by_name=None,
            registry_contact_person=normalize_text(values.get(36)),
            registry_exclusion_reason=normalize_text(values.get(4)),
            registry_update_plan_text=normalize_text(values.get(30)),
            row_payload=payload,
        )
        enrich_registry_row(row)
        rows.append(row)

    return rows


def load_deleted_rows(sheet) -> list[RegistryRow]:
    headers = {col_idx: sheet.cell(DELETED_HEADER_ROW, col_idx).value for col_idx in range(1, sheet.max_column + 1)}
    rows: list[RegistryRow] = []

    for row_num in range(DELETED_DATA_ROW, sheet.max_row + 1):
        values = {col_idx: sheet.cell(row_num, col_idx).value for col_idx in range(1, sheet.max_column + 1)}
        document_name = normalize_text(values.get(4))
        if not document_name:
            continue

        payload = make_payload(headers, values)
        row = RegistryRow(
            source_sheet=DELETED_SHEET,
            source_row_num=row_num,
            registry_scope="Удалено",
            core_registry_flag=False,
            registry_status_raw=None,
            registry_record_number=normalize_text(values.get(2)),
            registry_document_code=None,
            registry_document_name=document_name,
            registry_strategy_tactic=None,
            registry_ksu_flag=None,
            registry_business_process_id=None,
            registry_business_process_name=None,
            registry_document_type=None,
            registry_department_type=None,
            registry_department_name=normalize_text(values.get(3)),
            registry_department_short_name=None,
            registry_content_summary=None,
            registry_object_list=None,
            registry_stage_list=None,
            registry_document_about=None,
            registry_goal_text=None,
            registry_task_text=None,
            registry_project_link_text=None,
            registry_project_id_text=None,
            registry_project_link_alt_text=None,
            registry_project_id_alt_text=None,
            registry_incoming_documents=None,
            registry_impacted_documents=None,
            registry_input_data_text=None,
            registry_output_kpi_text=None,
            registry_trigger_documents=None,
            registry_information_recipients=None,
            registry_horizon_level_3=None,
            registry_horizon_strict=None,
            registry_period_label=None,
            registry_validity_text=None,
            registry_update_frequency_text=None,
            registry_actualization_date_raw=None,
            registry_status_text=None,
            registry_approved_by_position=None,
            registry_approved_by_name=None,
            registry_contact_person=None,
            registry_exclusion_reason=normalize_text(values.get(5)),
            registry_update_plan_text=None,
            row_payload=payload,
        )
        enrich_registry_row(row)
        rows.append(row)

    return rows


def enrich_registry_row(row: RegistryRow) -> None:
    row.canonical_department = canonical_department(row.registry_department_name, row.registry_department_short_name)
    row.match_name_norm = normalize_match_text(row.registry_document_name)
    row.token_set = token_set(row.registry_document_name)
    row.keyword_set = keyword_set(row.registry_document_name)


def create_connection():
    return psycopg2.connect(require_database_url())


def resolve_run_id(cursor, explicit_run_id: Optional[str]) -> str:
    if explicit_run_id:
        return explicit_run_id

    cursor.execute("SELECT run_id FROM mart.strategy_current_intake_run LIMIT 1")
    row = cursor.fetchone()
    if not row or not row[0]:
        raise RuntimeError("Current strategy intake run was not found")
    return row[0]


def fetch_ready_documents(cursor, run_id: str) -> list[IntakeDocument]:
    cursor.execute(
        """
        SELECT
            doc_id,
            document_short_name,
            document_full_name,
            department_code,
            horizon_code,
            status_code
        FROM staging.strategy_document_intake
        WHERE run_id = %s
            AND ready_for_dwh_flag = TRUE
        """,
        (run_id,),
    )
    rows = cursor.fetchall()
    if not rows:
        raise RuntimeError(f"No ready documents found for run_id={run_id}")

    documents: list[IntakeDocument] = []
    for row in rows:
        short_name = normalize_text(row[1]) or ""
        full_name = normalize_text(row[2])
        short_match = normalize_match_text(short_name)
        full_match = normalize_match_text(full_name)
        combined_match = " ".join(part for part in (short_match, full_match) if part).strip()
        documents.append(
            IntakeDocument(
                doc_id=row[0],
                short_name=short_name,
                full_name=full_name,
                department_code=normalize_text(row[3]),
                horizon_code=normalize_text(row[4]),
                status_code=normalize_text(row[5]),
                short_match=short_match,
                full_match=full_match,
                combined_match=combined_match,
                token_set=token_set(combined_match),
                keyword_set=keyword_set(combined_match),
            )
        )
    return documents


def overlap_ratio(left: frozenset[str], right: frozenset[str]) -> float:
    if not left or not right:
        return 0.0
    return len(left & right) / max(len(left), 1)


def horizon_bonus(registry_row: RegistryRow, document: IntakeDocument) -> float:
    row_horizon = canonical_horizon(registry_row.registry_horizon_level_3 or registry_row.registry_horizon_strict)
    if row_horizon and document.horizon_code and row_horizon == document.horizon_code:
        return 0.05
    return 0.0


def score_candidate(registry_row: RegistryRow, document: IntakeDocument) -> float:
    sequence_score = max(
        SequenceMatcher(None, registry_row.match_name_norm, document.short_match).ratio(),
        SequenceMatcher(None, registry_row.match_name_norm, document.full_match).ratio(),
        SequenceMatcher(None, registry_row.match_name_norm, document.combined_match).ratio(),
    )
    token_overlap = overlap_ratio(registry_row.token_set, document.token_set)
    keyword_overlap = overlap_ratio(registry_row.keyword_set, document.keyword_set)
    containment = (
        1.0
        if registry_row.match_name_norm
        and (
            registry_row.match_name_norm in document.combined_match
            or document.short_match in registry_row.match_name_norm
            or document.full_match in registry_row.match_name_norm
        )
        else 0.0
    )
    department_alignment = (
        0.06
        if registry_row.canonical_department
        and document.department_code
        and registry_row.canonical_department == document.department_code
        else 0.0
    )
    total = (
        (0.46 * sequence_score)
        + (0.24 * token_overlap)
        + (0.18 * keyword_overlap)
        + (0.07 * containment)
        + department_alignment
        + horizon_bonus(registry_row, document)
    )
    if registry_row.match_name_norm and registry_row.match_name_norm in {document.short_match, document.full_match}:
        total = max(total, 0.97)
    return round(min(total, 0.9999), 4)


def select_docs(
    documents: Iterable[IntakeDocument],
    include_all: Iterable[str] = (),
    include_any: Iterable[str] = (),
    exclude_any: Iterable[str] = (),
) -> list[IntakeDocument]:
    include_all_set = [normalize_match_text(term) for term in include_all if normalize_match_text(term)]
    include_any_set = [normalize_match_text(term) for term in include_any if normalize_match_text(term)]
    exclude_any_set = [normalize_match_text(term) for term in exclude_any if normalize_match_text(term)]

    selected: list[IntakeDocument] = []
    for document in documents:
        text = document.combined_match
        if include_all_set and not all(term in text for term in include_all_set):
            continue
        if include_any_set and not any(term in text for term in include_any_set):
            continue
        if exclude_any_set and any(term in text for term in exclude_any_set):
            continue
        selected.append(document)
    return selected


def low_signal_penalty(document: IntakeDocument) -> float:
    penalty = 0.0
    text = f"{document.short_match} {document.full_match}".strip()
    for pattern, pattern_penalty in LOW_SIGNAL_DOC_PATTERNS.items():
        normalized_pattern = normalize_match_text(pattern)
        if normalized_pattern and normalized_pattern in text:
            penalty = max(penalty, pattern_penalty)
    return penalty


def score_documents(registry_row: RegistryRow, documents: Iterable[IntakeDocument], min_score: float = 0.0):
    scored = []
    for document in documents:
        score = score_candidate(registry_row, document)
        if score >= min_score:
            scored.append((score, document))
    scored.sort(key=lambda item: (item[0], item[1].doc_id), reverse=True)
    return scored


def choose_primary(
    scored_documents,
    preferred_term_groups: Iterable[Iterable[str]] = (),
    demoted_terms: Iterable[str] = (),
):
    ranked = []
    normalized_preferred = [
        [normalize_match_text(term) for term in group if normalize_match_text(term)]
        for group in preferred_term_groups
    ]
    normalized_demoted = [normalize_match_text(term) for term in demoted_terms if normalize_match_text(term)]

    for base_score, document in scored_documents:
        adjusted_score = base_score
        for term_group in normalized_preferred:
            if term_group and all(term in document.combined_match for term in term_group):
                adjusted_score += 0.14
        for demoted_term in normalized_demoted:
            if demoted_term and demoted_term in document.combined_match:
                adjusted_score -= 0.10
        adjusted_score -= low_signal_penalty(document)
        ranked.append((round(adjusted_score, 4), base_score, document))

    ranked.sort(key=lambda item: (item[0], item[1], item[2].doc_id), reverse=True)
    return ranked


def make_doc_xwalk(
    registry_row: RegistryRow,
    source_system: str,
    source_snapshot_label: Optional[str],
    technical_snapshot_date: date,
    run_id: str,
    document: IntakeDocument,
    match_status: str,
    match_method: str,
    match_role: str,
    match_score: float,
    review_required_flag: bool,
    match_comment: Optional[str],
) -> XwalkRow:
    return XwalkRow(
        registry_raw_id=None,
        source_system=source_system,
        source_snapshot_label=source_snapshot_label,
        technical_snapshot_date=technical_snapshot_date,
        run_id=run_id,
        registry_scope=registry_row.registry_scope,
        match_status=match_status,
        match_method=match_method,
        match_role=match_role,
        candidate_rank=None,
        match_score=round(match_score, 4),
        review_required_flag=review_required_flag,
        doc_id=document.doc_id,
        intake_document_short_name=document.short_name,
        intake_document_full_name=document.full_name,
        intake_department_code=document.department_code,
        intake_horizon_code=document.horizon_code,
        intake_status_code=document.status_code,
        match_comment=match_comment,
    )


def make_empty_xwalk(
    registry_row: RegistryRow,
    source_system: str,
    source_snapshot_label: Optional[str],
    technical_snapshot_date: date,
    run_id: str,
    match_status: str,
    match_method: str,
    match_comment: Optional[str],
) -> XwalkRow:
    return XwalkRow(
        registry_raw_id=None,
        source_system=source_system,
        source_snapshot_label=source_snapshot_label,
        technical_snapshot_date=technical_snapshot_date,
        run_id=run_id,
        registry_scope=registry_row.registry_scope,
        match_status=match_status,
        match_method=match_method,
        match_role=MATCH_ROLE_NONE,
        candidate_rank=None,
        match_score=None,
        review_required_flag=False,
        doc_id=None,
        intake_document_short_name=None,
        intake_document_full_name=None,
        intake_department_code=None,
        intake_horizon_code=None,
        intake_status_code=None,
        match_comment=match_comment,
    )


def dedupe_and_rank(matches: list[XwalkRow]) -> list[XwalkRow]:
    deduped: dict[tuple[Any, ...], XwalkRow] = {}
    for match in matches:
        key = (match.doc_id, match.match_status, match.match_role, match.match_method, match.match_comment)
        existing = deduped.get(key)
        if existing is None or (match.match_score or 0.0) > (existing.match_score or 0.0):
            deduped[key] = match

    ordered = sorted(
        deduped.values(),
        key=lambda item: (
            0 if item.match_role == MATCH_ROLE_PRIMARY else 1 if item.match_role == MATCH_ROLE_COMPONENT else 2 if item.match_role == MATCH_ROLE_CANDIDATE else 3,
            -(item.match_score or 0.0),
            item.doc_id or "",
        ),
    )

    candidate_rank = 1
    for match in ordered:
        if match.doc_id:
            match.candidate_rank = candidate_rank
            candidate_rank += 1

    return ordered


def build_cluster_matches(
    registry_row: RegistryRow,
    source_system: str,
    source_snapshot_label: Optional[str],
    technical_snapshot_date: date,
    run_id: str,
    scored_documents,
    match_method: str,
    preferred_term_groups: Iterable[Iterable[str]],
    component_threshold: float = 0.48,
    review_required_flag: bool = False,
):
    if not scored_documents:
        return []

    ranked = choose_primary(
        scored_documents,
        preferred_term_groups=preferred_term_groups,
        demoted_terms=LOW_SIGNAL_DOC_PATTERNS.keys(),
    )
    if not ranked:
        return []

    primary_adjusted, primary_base, primary_document = ranked[0]
    if primary_base < 0.45:
        return []

    matches = [
        make_doc_xwalk(
            registry_row,
            source_system,
            source_snapshot_label,
            technical_snapshot_date,
            run_id,
            primary_document,
            MATCH_STATUS_AUTO,
            match_method,
            MATCH_ROLE_PRIMARY,
            primary_base,
            review_required_flag,
            f"Кластерное сопоставление; выбран основной документ из {len(ranked)} вариантов.",
        )
    ]

    for adjusted_score, base_score, document in ranked[1:]:
        if base_score < component_threshold:
            continue
        matches.append(
            make_doc_xwalk(
                registry_row,
                source_system,
                source_snapshot_label,
                technical_snapshot_date,
                run_id,
                document,
                MATCH_STATUS_AUTO,
                match_method,
                MATCH_ROLE_COMPONENT,
                base_score,
                review_required_flag,
                "Компонент того же документного кластера.",
            )
        )

    return dedupe_and_rank(matches)


def build_primary_with_candidates(
    registry_row: RegistryRow,
    source_system: str,
    source_snapshot_label: Optional[str],
    technical_snapshot_date: date,
    run_id: str,
    scored_documents,
    match_method: str,
    preferred_term_groups: Iterable[Iterable[str]] = (),
    demoted_terms: Iterable[str] = (),
    candidate_threshold: float = 0.52,
    review_required_flag: bool = True,
):
    if not scored_documents:
        return []

    ranked = choose_primary(scored_documents, preferred_term_groups=preferred_term_groups, demoted_terms=demoted_terms)
    if not ranked:
        return []

    primary_adjusted, primary_base, primary_document = ranked[0]
    if primary_base < 0.45:
        return []

    matches = [
        make_doc_xwalk(
            registry_row,
            source_system,
            source_snapshot_label,
            technical_snapshot_date,
            run_id,
            primary_document,
            MATCH_STATUS_AUTO,
            match_method,
            MATCH_ROLE_PRIMARY,
            primary_base,
            review_required_flag,
            "Основной документ выбран автоматически; альтернативы сохранены как кандидаты.",
        )
    ]

    for adjusted_score, base_score, document in ranked[1:]:
        if base_score < candidate_threshold:
            continue
        matches.append(
            make_doc_xwalk(
                registry_row,
                source_system,
                source_snapshot_label,
                technical_snapshot_date,
                run_id,
                document,
                MATCH_STATUS_AMBIGUOUS,
                match_method,
                MATCH_ROLE_CANDIDATE,
                base_score,
                True,
                "Альтернативная версия или близкий документ; нужна проверка.",
            )
        )

    return dedupe_and_rank(matches)


def build_candidate_review(
    registry_row: RegistryRow,
    source_system: str,
    source_snapshot_label: Optional[str],
    technical_snapshot_date: date,
    run_id: str,
    scored_documents,
    match_method: str,
    limit: int = 3,
):
    matches = []
    for score, document in scored_documents[:limit]:
        matches.append(
            make_doc_xwalk(
                registry_row,
                source_system,
                source_snapshot_label,
                technical_snapshot_date,
                run_id,
                document,
                MATCH_STATUS_AMBIGUOUS,
                match_method,
                MATCH_ROLE_CANDIDATE,
                score,
                True,
                "Есть близкий кандидат, но автосопоставление недостаточно надежно.",
            )
        )
    return dedupe_and_rank(matches)


def special_matches(
    registry_row: RegistryRow,
    source_system: str,
    source_snapshot_label: Optional[str],
    technical_snapshot_date: date,
    run_id: str,
    pool: list[IntakeDocument],
):
    name = registry_row.match_name_norm

    if "дтср" in name:
        documents = select_docs(pool, include_any=("дтср",))
        scored = score_documents(registry_row, documents, min_score=0.34)
        return build_cluster_matches(
            registry_row,
            source_system,
            source_snapshot_label,
            technical_snapshot_date,
            run_id,
            scored,
            match_method="Кластер ДТСР",
            preferred_term_groups=(("книга", "1"), ("текст",)),
            component_threshold=0.46,
            review_required_flag=False,
        )

    if "тпр" in name and "лун" in name:
        documents = select_docs(pool, include_all=("тпр", "лун"))
        scored = score_documents(registry_row, documents, min_score=0.34)
        return build_cluster_matches(
            registry_row,
            source_system,
            source_snapshot_label,
            technical_snapshot_date,
            run_id,
            scored,
            match_method="Кластер ТПР Лунского",
            preferred_term_groups=(("книга", "1"), ("текст",)),
            component_threshold=0.46,
            review_required_flag=False,
        )

    if "укпо" in name or "уточненный данные к комплексному плану освоения" in name or "уточненный данные" in name:
        documents = select_docs(pool, include_any=("укпо",))
        scored = score_documents(registry_row, documents, min_score=0.48)
        return build_primary_with_candidates(
            registry_row,
            source_system,
            source_snapshot_label,
            technical_snapshot_date,
            run_id,
            scored,
            match_method="Версионное сопоставление УКПО",
            preferred_term_groups=(("2013",),),
            demoted_terms=("проект", "анг"),
            candidate_threshold=0.52,
            review_required_flag=True,
        )

    if "долгосрочный прогноз добычи" in name:
        documents = select_docs(pool, include_any=("ltf", "дпд"))
        scored = score_documents(registry_row, documents, min_score=0.30)
        ranked = choose_primary(
            scored,
            preferred_term_groups=(("ltf",), ("дпд",)),
            demoted_terms=("таблица", "приложение"),
        )
        if ranked:
            _, primary_base, primary_document = ranked[0]
            if primary_base >= 0.38:
                matches = [
                    make_doc_xwalk(
                        registry_row,
                        source_system,
                        source_snapshot_label,
                        technical_snapshot_date,
                        run_id,
                        primary_document,
                        MATCH_STATUS_AUTO,
                        "Сопоставление долгосрочного прогноза добычи",
                        MATCH_ROLE_PRIMARY,
                        primary_base,
                        False,
                        "Выбран основной LTF/ДПД-документ в узком кластере долгосрочного прогноза добычи.",
                    )
                ]
                for _, candidate_base, candidate_document in ranked[1:]:
                    if candidate_base >= 0.50:
                        matches.append(
                            make_doc_xwalk(
                                registry_row,
                                source_system,
                                source_snapshot_label,
                                technical_snapshot_date,
                                run_id,
                                candidate_document,
                                MATCH_STATUS_AMBIGUOUS,
                                "Сопоставление долгосрочного прогноза добычи",
                                MATCH_ROLE_CANDIDATE,
                                candidate_base,
                                True,
                                "Побочный артефакт того же LTF-кластера; оставлен кандидатом.",
                            )
                        )
                return dedupe_and_rank(matches)

    if "буровых работ" in name and "краткоср" in name:
        documents = select_docs(pool, include_any=("буровых",))
        scored = score_documents(registry_row, documents, min_score=0.50)
        if scored:
            top_score, top_document = scored[0]
            return dedupe_and_rank(
                [
                    make_doc_xwalk(
                        registry_row,
                        source_system,
                        source_snapshot_label,
                        technical_snapshot_date,
                        run_id,
                        top_document,
                        MATCH_STATUS_AUTO,
                        "Сопоставление программы буровых работ",
                        MATCH_ROLE_PRIMARY,
                        top_score,
                        False,
                        "Совпадает краткосрочный буровой документ.",
                    )
                ]
            )

    if "буровых работ" in name and "долгоср" in name:
        documents = select_docs(pool, include_any=("буровых",))
        scored = score_documents(registry_row, documents, min_score=0.42)
        return build_candidate_review(
            registry_row,
            source_system,
            source_snapshot_label,
            technical_snapshot_date,
            run_id,
            scored,
            match_method="Кандидат по буровым работам",
            limit=2,
        )

    if "морские объект" in name and "2 года" in name:
        documents = select_docs(pool, include_all=("кпд", "морской"), include_any=("2 года",))
        if not documents:
            documents = select_docs(pool, include_all=("кпд", "морской"), include_any=("2",))
        if not documents:
            documents = select_docs(pool, include_all=("кпд", "морской"))
        scored = score_documents(registry_row, documents, min_score=0.42)
        return build_primary_with_candidates(
            registry_row,
            source_system,
            source_snapshot_label,
            technical_snapshot_date,
            run_id,
            scored,
            match_method="Сопоставление КПД морских объектов",
            preferred_term_groups=(("2",), ("краткосрочный",)),
            demoted_terms=(),
            candidate_threshold=0.48,
            review_required_flag=False,
        )

    if "морские объект" in name and "5 лет" in name:
        documents = select_docs(pool, include_all=("кпд", "морской"), include_any=("5", "летний"))
        if not documents:
            documents = select_docs(pool, include_all=("кпд", "морской"))
        scored = score_documents(registry_row, documents, min_score=0.42)
        return build_primary_with_candidates(
            registry_row,
            source_system,
            source_snapshot_label,
            technical_snapshot_date,
            run_id,
            scored,
            match_method="Сопоставление КПД морских объектов",
            preferred_term_groups=(("5",), ("летний",)),
            demoted_terms=(),
            candidate_threshold=0.48,
            review_required_flag=False,
        )

    if "техническому обслуживанию" in name or "технического обслуживания" in name:
        dpto_map = [
            (("па а",), "ДПТО ПА-А"),
            (("па б",), "ДПТО ПА-Б"),
            (("лун а",), "ДПТО ЛУН-А"),
            (("нкс 2",), "ДПТО НКС-2"),
            (("уоуг",), "ДПТО нефтегазопроводы и УОУГ"),
            (("нефтегазопровод",), "ДПТО нефтегазопроводы и УОУГ"),
            (("обтк",), "ДПТО ОБТК"),
            (("спг тон",), "ДПТО СПГ/ТОН/ВПУ"),
        ]
        for terms, method_label in dpto_map:
            if all(term in name for term in terms):
                documents = select_docs(pool, include_all=("дпто", *terms))
                if not documents and "уоуг" in terms:
                    documents = select_docs(pool, include_all=("дпто",), include_any=("уоуг", "нефтегазопровод"))
                scored = score_documents(registry_row, documents, min_score=0.48)
                return build_primary_with_candidates(
                    registry_row,
                    source_system,
                    source_snapshot_label,
                    technical_snapshot_date,
                    run_id,
                    scored,
                    match_method=method_label,
                    preferred_term_groups=(terms,),
                    demoted_terms=(),
                    candidate_threshold=0.52,
                    review_required_flag=False,
                )

    if "маркетинг" in name and "спг" in name:
        documents = select_docs(pool, include_all=("маркетинг", "спг"))
        scored = score_documents(registry_row, documents, min_score=0.48)
        return build_primary_with_candidates(
            registry_row,
            source_system,
            source_snapshot_label,
            technical_snapshot_date,
            run_id,
            scored,
            match_method="План маркетинга СПГ",
            preferred_term_groups=(("маркетинг", "спг"),),
            demoted_terms=(),
            candidate_threshold=0.52,
            review_required_flag=False,
        )

    if "газовоз" in name:
        documents = select_docs(pool, include_any=("газовоз",))
        scored = score_documents(registry_row, documents, min_score=0.48)
        return build_primary_with_candidates(
            registry_row,
            source_system,
            source_snapshot_label,
            technical_snapshot_date,
            run_id,
            scored,
            match_method="План обеспечения газовозами",
            preferred_term_groups=(("газовоз",),),
            demoted_terms=(),
            candidate_threshold=0.50,
            review_required_flag=False,
        )

    if "танкер" in name or "нефтеналивный" in name:
        documents = select_docs(pool, include_any=("танкер",))
        scored = score_documents(registry_row, documents, min_score=0.46)
        return build_primary_with_candidates(
            registry_row,
            source_system,
            source_snapshot_label,
            technical_snapshot_date,
            run_id,
            scored,
            match_method="План обеспечения танкерами",
            preferred_term_groups=(("танкер",),),
            demoted_terms=("газовоз",),
            candidate_threshold=0.48,
            review_required_flag=True,
        )

    if "поставка" in name and "спг" in name:
        documents = select_docs(pool, include_all=("поставка", "спг"))
        scored = score_documents(registry_row, documents, min_score=0.48)
        return build_primary_with_candidates(
            registry_row,
            source_system,
            source_snapshot_label,
            technical_snapshot_date,
            run_id,
            scored,
            match_method="Программа поставок СПГ",
            preferred_term_groups=(("поставка", "спг"),),
            demoted_terms=(),
            candidate_threshold=0.50,
            review_required_flag=False,
        )

    if "климатическ" in name:
        documents = select_docs(pool, include_any=("климатическ",))
        scored = score_documents(registry_row, documents, min_score=0.48)
        return build_primary_with_candidates(
            registry_row,
            source_system,
            source_snapshot_label,
            technical_snapshot_date,
            run_id,
            scored,
            match_method="Климатическая стратегия",
            preferred_term_groups=(("климатическ",),),
            demoted_terms=("сценарий", "бенчмаркинг"),
            candidate_threshold=0.60,
            review_required_flag=True,
        )

    if "esg" in name:
        documents = select_docs(pool, include_any=("esg",))
        scored = score_documents(registry_row, documents, min_score=0.40)
        return build_candidate_review(
            registry_row,
            source_system,
            source_snapshot_label,
            technical_snapshot_date,
            run_id,
            scored,
            match_method="Кандидат ESG",
            limit=2,
        )

    if "изменение" in name and "смете расходов на освоение" in name:
        documents = select_docs(pool, include_any=("исо",))
        scored = score_documents(registry_row, documents, min_score=0.40)
        return build_primary_with_candidates(
            registry_row,
            source_system,
            source_snapshot_label,
            technical_snapshot_date,
            run_id,
            scored,
            match_method="Сопоставление ИСО",
            preferred_term_groups=(("исо",),),
            demoted_terms=(),
            candidate_threshold=0.44,
            review_required_flag=True,
        )

    if "программа работ и смета расходов" in name:
        documents = select_docs(pool, include_any=("пр и ср", "смета расходов"))
        scored = score_documents(registry_row, documents, min_score=0.46)
        return build_primary_with_candidates(
            registry_row,
            source_system,
            source_snapshot_label,
            technical_snapshot_date,
            run_id,
            scored,
            match_method="Сопоставление ПР и СР",
            preferred_term_groups=(("пр", "ср"), ("смета", "расходов")),
            demoted_terms=(),
            candidate_threshold=0.48,
            review_required_flag=False,
        )

    if "стратегический план управления объектами" in name:
        documents = select_docs(pool, include_all=("управления", "объект"))
        scored = score_documents(registry_row, documents, min_score=0.48)
        return build_primary_with_candidates(
            registry_row,
            source_system,
            source_snapshot_label,
            technical_snapshot_date,
            run_id,
            scored,
            match_method="План управления объектами",
            preferred_term_groups=(("управления", "объект"),),
            demoted_terms=(),
            candidate_threshold=0.54,
            review_required_flag="_new" in (registry_row.registry_document_name or ""),
        )

    if "стратегия развития общества" in name:
        documents = select_docs(pool, include_all=("стратегия", "общества"))
        scored = score_documents(registry_row, documents, min_score=0.40)
        return build_primary_with_candidates(
            registry_row,
            source_system,
            source_snapshot_label,
            technical_snapshot_date,
            run_id,
            scored,
            match_method="Стратегия Общества",
            preferred_term_groups=(("стратегия", "общества"),),
            demoted_terms=(),
            candidate_threshold=0.44,
            review_required_flag=True,
        )

    if "кадровый директорат" in name and "стратегия" in name:
        documents = select_docs(pool, include_any=("кадровый стратегия", "кадровый"))
        scored = score_documents(registry_row, documents, min_score=0.40)
        return build_primary_with_candidates(
            registry_row,
            source_system,
            source_snapshot_label,
            technical_snapshot_date,
            run_id,
            scored,
            match_method="Стратегия кадрового направления",
            preferred_term_groups=(("кадровый", "стратегия"),),
            demoted_terms=(),
            candidate_threshold=0.46,
            review_required_flag=True,
        )

    return []


def general_matches(
    registry_row: RegistryRow,
    source_system: str,
    source_snapshot_label: Optional[str],
    technical_snapshot_date: date,
    run_id: str,
    pool: list[IntakeDocument],
):
    scored = score_documents(registry_row, pool, min_score=0.40)
    if not scored:
        return []

    exact_docs = [
        document
        for document in pool
        if registry_row.match_name_norm and registry_row.match_name_norm in {document.short_match, document.full_match}
    ]
    if len(exact_docs) == 1:
        exact_document = exact_docs[0]
        return dedupe_and_rank(
            [
                make_doc_xwalk(
                    registry_row,
                    source_system,
                    source_snapshot_label,
                    technical_snapshot_date,
                    run_id,
                    exact_document,
                    MATCH_STATUS_AUTO,
                    "Строгое совпадение названия",
                    MATCH_ROLE_PRIMARY,
                    max(score_candidate(registry_row, exact_document), 0.97),
                    False,
                    "Строгое совпадение названия документа в пределах подразделения.",
                )
            ]
        )

    top_score, top_document = scored[0]
    second_score = scored[1][0] if len(scored) > 1 else 0.0

    if top_score >= 0.88 or (top_score >= 0.74 and (top_score - second_score) >= 0.08):
        return dedupe_and_rank(
            [
                make_doc_xwalk(
                    registry_row,
                    source_system,
                    source_snapshot_label,
                    technical_snapshot_date,
                    run_id,
                    top_document,
                    MATCH_STATUS_AUTO,
                    "Автосопоставление по близости названия",
                    MATCH_ROLE_PRIMARY,
                    top_score,
                    False,
                    "Лучший кандидат уверенно опережает остальные по близости названия.",
                )
            ]
        )

    if top_score >= 0.62:
        review_pool = [item for item in scored if item[0] >= top_score - 0.08]
        return build_candidate_review(
            registry_row,
            source_system,
            source_snapshot_label,
            technical_snapshot_date,
            run_id,
            review_pool,
            match_method="Нужна ручная проверка близких названий",
            limit=3,
        )

    if top_score >= 0.52:
        return build_candidate_review(
            registry_row,
            source_system,
            source_snapshot_label,
            technical_snapshot_date,
            run_id,
            scored,
            match_method="Слабый кандидат по близости названия",
            limit=2,
        )

    return []


def match_registry_row(
    registry_row: RegistryRow,
    source_system: str,
    source_snapshot_label: Optional[str],
    technical_snapshot_date: date,
    run_id: str,
    ready_documents: list[IntakeDocument],
) -> list[XwalkRow]:
    if registry_row.canonical_department:
        department_pool = [doc for doc in ready_documents if doc.department_code == registry_row.canonical_department]
    else:
        department_pool = list(ready_documents)

    matches = special_matches(
        registry_row,
        source_system,
        source_snapshot_label,
        technical_snapshot_date,
        run_id,
        department_pool,
    )
    if matches:
        return matches

    matches = general_matches(
        registry_row,
        source_system,
        source_snapshot_label,
        technical_snapshot_date,
        run_id,
        department_pool,
    )
    if matches:
        return matches

    if registry_row.registry_scope != "База":
        return [
            make_empty_xwalk(
                registry_row,
                source_system,
                source_snapshot_label,
                technical_snapshot_date,
                run_id,
                MATCH_STATUS_REJECTED,
                "Вне базового контура Excel",
                registry_row.registry_exclusion_reason or "Строка относится к исключенному или удаленному слою реестра.",
            )
        ]

    return [
        make_empty_xwalk(
            registry_row,
            source_system,
            source_snapshot_label,
            technical_snapshot_date,
            run_id,
            MATCH_STATUS_UNMATCHED,
            "Автосопоставление не найдено",
            "Для строки не найден надежный intake-документ в текущем ready-контуре.",
        )
    ]


def outcome_from_matches(matches: list[XwalkRow]) -> str:
    if any(match.doc_id and match.match_role == MATCH_ROLE_PRIMARY and match.review_required_flag for match in matches):
        return OUTCOME_MATCHED_REVIEW
    if any(match.doc_id and match.match_role == MATCH_ROLE_PRIMARY for match in matches):
        return OUTCOME_MATCHED
    if any(match.doc_id and match.match_role == MATCH_ROLE_CANDIDATE for match in matches):
        return OUTCOME_REVIEW
    if any(match.match_status == MATCH_STATUS_REJECTED for match in matches):
        return OUTCOME_REJECTED
    return OUTCOME_UNMATCHED


RAW_UPSERT_SQL = """
INSERT INTO staging.strategy_registry_overlay_raw (
    source_system,
    source_snapshot_label,
    technical_snapshot_date,
    source_file,
    source_sheet,
    source_row_num,
    registry_scope,
    core_registry_flag,
    registry_status_raw,
    registry_record_number,
    registry_document_code,
    registry_document_name,
    registry_strategy_tactic,
    registry_ksu_flag,
    registry_business_process_id,
    registry_business_process_name,
    registry_document_type,
    registry_department_type,
    registry_department_name,
    registry_department_short_name,
    registry_content_summary,
    registry_object_list,
    registry_stage_list,
    registry_document_about,
    registry_goal_text,
    registry_task_text,
    registry_project_link_text,
    registry_project_id_text,
    registry_project_link_alt_text,
    registry_project_id_alt_text,
    registry_incoming_documents,
    registry_impacted_documents,
    registry_input_data_text,
    registry_output_kpi_text,
    registry_trigger_documents,
    registry_information_recipients,
    registry_horizon_level_3,
    registry_horizon_strict,
    registry_period_label,
    registry_validity_text,
    registry_update_frequency_text,
    registry_actualization_date_raw,
    registry_status_text,
    registry_approved_by_position,
    registry_approved_by_name,
    registry_contact_person,
    registry_exclusion_reason,
    registry_update_plan_text,
    row_payload
) VALUES (
    %(source_system)s,
    %(source_snapshot_label)s,
    %(technical_snapshot_date)s,
    %(source_file)s,
    %(source_sheet)s,
    %(source_row_num)s,
    %(registry_scope)s,
    %(core_registry_flag)s,
    %(registry_status_raw)s,
    %(registry_record_number)s,
    %(registry_document_code)s,
    %(registry_document_name)s,
    %(registry_strategy_tactic)s,
    %(registry_ksu_flag)s,
    %(registry_business_process_id)s,
    %(registry_business_process_name)s,
    %(registry_document_type)s,
    %(registry_department_type)s,
    %(registry_department_name)s,
    %(registry_department_short_name)s,
    %(registry_content_summary)s,
    %(registry_object_list)s,
    %(registry_stage_list)s,
    %(registry_document_about)s,
    %(registry_goal_text)s,
    %(registry_task_text)s,
    %(registry_project_link_text)s,
    %(registry_project_id_text)s,
    %(registry_project_link_alt_text)s,
    %(registry_project_id_alt_text)s,
    %(registry_incoming_documents)s,
    %(registry_impacted_documents)s,
    %(registry_input_data_text)s,
    %(registry_output_kpi_text)s,
    %(registry_trigger_documents)s,
    %(registry_information_recipients)s,
    %(registry_horizon_level_3)s,
    %(registry_horizon_strict)s,
    %(registry_period_label)s,
    %(registry_validity_text)s,
    %(registry_update_frequency_text)s,
    %(registry_actualization_date_raw)s,
    %(registry_status_text)s,
    %(registry_approved_by_position)s,
    %(registry_approved_by_name)s,
    %(registry_contact_person)s,
    %(registry_exclusion_reason)s,
    %(registry_update_plan_text)s,
    %(row_payload)s
)
ON CONFLICT (source_system, technical_snapshot_date, source_sheet, source_row_num) DO UPDATE SET
    source_snapshot_label = EXCLUDED.source_snapshot_label,
    source_file = EXCLUDED.source_file,
    registry_scope = EXCLUDED.registry_scope,
    core_registry_flag = EXCLUDED.core_registry_flag,
    registry_status_raw = EXCLUDED.registry_status_raw,
    registry_record_number = EXCLUDED.registry_record_number,
    registry_document_code = EXCLUDED.registry_document_code,
    registry_document_name = EXCLUDED.registry_document_name,
    registry_strategy_tactic = EXCLUDED.registry_strategy_tactic,
    registry_ksu_flag = EXCLUDED.registry_ksu_flag,
    registry_business_process_id = EXCLUDED.registry_business_process_id,
    registry_business_process_name = EXCLUDED.registry_business_process_name,
    registry_document_type = EXCLUDED.registry_document_type,
    registry_department_type = EXCLUDED.registry_department_type,
    registry_department_name = EXCLUDED.registry_department_name,
    registry_department_short_name = EXCLUDED.registry_department_short_name,
    registry_content_summary = EXCLUDED.registry_content_summary,
    registry_object_list = EXCLUDED.registry_object_list,
    registry_stage_list = EXCLUDED.registry_stage_list,
    registry_document_about = EXCLUDED.registry_document_about,
    registry_goal_text = EXCLUDED.registry_goal_text,
    registry_task_text = EXCLUDED.registry_task_text,
    registry_project_link_text = EXCLUDED.registry_project_link_text,
    registry_project_id_text = EXCLUDED.registry_project_id_text,
    registry_project_link_alt_text = EXCLUDED.registry_project_link_alt_text,
    registry_project_id_alt_text = EXCLUDED.registry_project_id_alt_text,
    registry_incoming_documents = EXCLUDED.registry_incoming_documents,
    registry_impacted_documents = EXCLUDED.registry_impacted_documents,
    registry_input_data_text = EXCLUDED.registry_input_data_text,
    registry_output_kpi_text = EXCLUDED.registry_output_kpi_text,
    registry_trigger_documents = EXCLUDED.registry_trigger_documents,
    registry_information_recipients = EXCLUDED.registry_information_recipients,
    registry_horizon_level_3 = EXCLUDED.registry_horizon_level_3,
    registry_horizon_strict = EXCLUDED.registry_horizon_strict,
    registry_period_label = EXCLUDED.registry_period_label,
    registry_validity_text = EXCLUDED.registry_validity_text,
    registry_update_frequency_text = EXCLUDED.registry_update_frequency_text,
    registry_actualization_date_raw = EXCLUDED.registry_actualization_date_raw,
    registry_status_text = EXCLUDED.registry_status_text,
    registry_approved_by_position = EXCLUDED.registry_approved_by_position,
    registry_approved_by_name = EXCLUDED.registry_approved_by_name,
    registry_contact_person = EXCLUDED.registry_contact_person,
    registry_exclusion_reason = EXCLUDED.registry_exclusion_reason,
    registry_update_plan_text = EXCLUDED.registry_update_plan_text,
    row_payload = EXCLUDED.row_payload
RETURNING registry_raw_id
"""


XWALK_INSERT_SQL = """
INSERT INTO staging.strategy_registry_overlay_xwalk (
    registry_raw_id,
    source_system,
    source_snapshot_label,
    technical_snapshot_date,
    run_id,
    record_source,
    registry_scope,
    match_status,
    match_method,
    match_role,
    candidate_rank,
    match_score,
    review_required_flag,
    doc_id,
    intake_document_short_name,
    intake_document_full_name,
    intake_department_code,
    intake_horizon_code,
    intake_status_code,
    match_comment
) VALUES (
    %(registry_raw_id)s,
    %(source_system)s,
    %(source_snapshot_label)s,
    %(technical_snapshot_date)s,
    %(run_id)s,
    %(record_source)s,
    %(registry_scope)s,
    %(match_status)s,
    %(match_method)s,
    %(match_role)s,
    %(candidate_rank)s,
    %(match_score)s,
    %(review_required_flag)s,
    %(doc_id)s,
    %(intake_document_short_name)s,
    %(intake_document_full_name)s,
    %(intake_department_code)s,
    %(intake_horizon_code)s,
    %(intake_status_code)s,
    %(match_comment)s
)
"""


def upsert_raw_rows(
    cursor,
    registry_rows: list[RegistryRow],
    source_system: str,
    source_snapshot_label: Optional[str],
    technical_snapshot_date: date,
    source_file: str,
) -> dict[tuple[str, int], str]:
    row_id_map: dict[tuple[str, int], str] = {}
    for registry_row in registry_rows:
        payload = {
            "source_system": source_system,
            "source_snapshot_label": source_snapshot_label,
            "technical_snapshot_date": technical_snapshot_date,
            "source_file": source_file,
            "source_sheet": registry_row.source_sheet,
            "source_row_num": registry_row.source_row_num,
            "registry_scope": registry_row.registry_scope,
            "core_registry_flag": registry_row.core_registry_flag,
            "registry_status_raw": registry_row.registry_status_raw,
            "registry_record_number": registry_row.registry_record_number,
            "registry_document_code": registry_row.registry_document_code,
            "registry_document_name": registry_row.registry_document_name,
            "registry_strategy_tactic": registry_row.registry_strategy_tactic,
            "registry_ksu_flag": registry_row.registry_ksu_flag,
            "registry_business_process_id": registry_row.registry_business_process_id,
            "registry_business_process_name": registry_row.registry_business_process_name,
            "registry_document_type": registry_row.registry_document_type,
            "registry_department_type": registry_row.registry_department_type,
            "registry_department_name": registry_row.registry_department_name,
            "registry_department_short_name": registry_row.registry_department_short_name,
            "registry_content_summary": registry_row.registry_content_summary,
            "registry_object_list": registry_row.registry_object_list,
            "registry_stage_list": registry_row.registry_stage_list,
            "registry_document_about": registry_row.registry_document_about,
            "registry_goal_text": registry_row.registry_goal_text,
            "registry_task_text": registry_row.registry_task_text,
            "registry_project_link_text": registry_row.registry_project_link_text,
            "registry_project_id_text": registry_row.registry_project_id_text,
            "registry_project_link_alt_text": registry_row.registry_project_link_alt_text,
            "registry_project_id_alt_text": registry_row.registry_project_id_alt_text,
            "registry_incoming_documents": registry_row.registry_incoming_documents,
            "registry_impacted_documents": registry_row.registry_impacted_documents,
            "registry_input_data_text": registry_row.registry_input_data_text,
            "registry_output_kpi_text": registry_row.registry_output_kpi_text,
            "registry_trigger_documents": registry_row.registry_trigger_documents,
            "registry_information_recipients": registry_row.registry_information_recipients,
            "registry_horizon_level_3": registry_row.registry_horizon_level_3,
            "registry_horizon_strict": registry_row.registry_horizon_strict,
            "registry_period_label": registry_row.registry_period_label,
            "registry_validity_text": registry_row.registry_validity_text,
            "registry_update_frequency_text": registry_row.registry_update_frequency_text,
            "registry_actualization_date_raw": registry_row.registry_actualization_date_raw,
            "registry_status_text": registry_row.registry_status_text,
            "registry_approved_by_position": registry_row.registry_approved_by_position,
            "registry_approved_by_name": registry_row.registry_approved_by_name,
            "registry_contact_person": registry_row.registry_contact_person,
            "registry_exclusion_reason": registry_row.registry_exclusion_reason,
            "registry_update_plan_text": registry_row.registry_update_plan_text,
            "row_payload": Json(registry_row.row_payload),
        }
        cursor.execute(RAW_UPSERT_SQL, payload)
        registry_raw_id = cursor.fetchone()[0]
        row_id_map[(registry_row.source_sheet, registry_row.source_row_num)] = registry_raw_id

    return row_id_map


def refresh_auto_xwalk(
    cursor,
    xwalk_rows: list[XwalkRow],
    source_system: str,
    technical_snapshot_date: date,
    run_id: str,
) -> None:
    cursor.execute(
        """
        DELETE FROM staging.strategy_registry_overlay_xwalk
        WHERE source_system = %s
            AND technical_snapshot_date = %s
            AND run_id = %s
            AND record_source = %s
        """,
        (source_system, technical_snapshot_date, run_id, RECORD_SOURCE_AUTO),
    )

    for xwalk_row in xwalk_rows:
        cursor.execute(
            XWALK_INSERT_SQL,
            {
                "registry_raw_id": xwalk_row.registry_raw_id,
                "source_system": xwalk_row.source_system,
                "source_snapshot_label": xwalk_row.source_snapshot_label,
                "technical_snapshot_date": xwalk_row.technical_snapshot_date,
                "run_id": xwalk_row.run_id,
                "record_source": xwalk_row.record_source,
                "registry_scope": xwalk_row.registry_scope,
                "match_status": xwalk_row.match_status,
                "match_method": xwalk_row.match_method,
                "match_role": xwalk_row.match_role,
                "candidate_rank": xwalk_row.candidate_rank,
                "match_score": xwalk_row.match_score,
                "review_required_flag": xwalk_row.review_required_flag,
                "doc_id": xwalk_row.doc_id,
                "intake_document_short_name": xwalk_row.intake_document_short_name,
                "intake_document_full_name": xwalk_row.intake_document_full_name,
                "intake_department_code": xwalk_row.intake_department_code,
                "intake_horizon_code": xwalk_row.intake_horizon_code,
                "intake_status_code": xwalk_row.intake_status_code,
                "match_comment": xwalk_row.match_comment,
            },
        )


def summarize(registry_rows: list[RegistryRow], row_matches: list[tuple[RegistryRow, list[XwalkRow]]]) -> str:
    total_rows = len(registry_rows)
    scope_counts: dict[str, int] = {}
    outcome_counts: dict[str, int] = {}
    unmatched_names: list[str] = []

    for registry_row, matches in row_matches:
        scope_counts[registry_row.registry_scope] = scope_counts.get(registry_row.registry_scope, 0) + 1
        if registry_row.registry_scope == "База":
            outcome = outcome_from_matches(matches)
            outcome_counts[outcome] = outcome_counts.get(outcome, 0) + 1
            if outcome == OUTCOME_UNMATCHED:
                unmatched_names.append(registry_row.registry_document_name)

    parts = [
        f"Всего строк реестра: {total_rows}",
        "По слоям: " + ", ".join(f"{scope}={count}" for scope, count in sorted(scope_counts.items())),
        "Базовый контур: "
        + ", ".join(f"{status}={count}" for status, count in sorted(outcome_counts.items())),
    ]
    if unmatched_names:
        parts.append("База без автосопоставления: " + "; ".join(unmatched_names[:12]))
    return "\n".join(parts)


def verify_loaded_state(cursor, source_system: str, technical_snapshot_date: date, run_id: str) -> str:
    cursor.execute(
        """
        SELECT registry_scope, COUNT(*)
        FROM staging.strategy_registry_overlay_raw
        WHERE source_system = %s
            AND technical_snapshot_date = %s
        GROUP BY registry_scope
        ORDER BY registry_scope
        """,
        (source_system, technical_snapshot_date),
    )
    raw_counts = cursor.fetchall()

    cursor.execute(
        """
        SELECT match_status, COUNT(*)
        FROM staging.strategy_registry_overlay_xwalk
        WHERE source_system = %s
            AND technical_snapshot_date = %s
            AND run_id = %s
        GROUP BY match_status
        ORDER BY match_status
        """,
        (source_system, technical_snapshot_date, run_id),
    )
    xwalk_counts = cursor.fetchall()

    cursor.execute(
        """
        SELECT xwalk_status, COUNT(*)
        FROM mart.strategy_core_documents
        WHERE source_system = %s
            AND technical_snapshot_date = %s
        GROUP BY xwalk_status
        ORDER BY xwalk_status
        """,
        (source_system, technical_snapshot_date),
    )
    mart_counts = cursor.fetchall()

    return "\n".join(
        [
            "staging.strategy_registry_overlay_raw: "
            + ", ".join(f"{scope}={count}" for scope, count in raw_counts),
            "staging.strategy_registry_overlay_xwalk: "
            + ", ".join(f"{status}={count}" for status, count in xwalk_counts),
            "mart.strategy_core_documents: "
            + ", ".join(f"{status}={count}" for status, count in mart_counts),
        ]
    )


def main(argv: list[str]) -> int:
    args = parse_args(argv)

    workbook_path = Path(args.workbook).expanduser().resolve()
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    technical_snapshot_date = (
        datetime.strptime(args.technical_snapshot_date, "%Y-%m-%d").date()
        if args.technical_snapshot_date
        else datetime.fromtimestamp(workbook_path.stat().st_mtime).date()
    )
    source_system = args.source_system
    source_snapshot_label = args.source_snapshot_label or workbook_path.name

    registry_rows = load_workbook_rows(workbook_path)

    with create_connection() as connection:
        with connection.cursor() as cursor:
            run_id = resolve_run_id(cursor, args.run_id)
            ready_documents = fetch_ready_documents(cursor, run_id)

            matched_rows: list[tuple[RegistryRow, list[XwalkRow]]] = []
            for registry_row in registry_rows:
                matches = match_registry_row(
                    registry_row,
                    source_system,
                    source_snapshot_label,
                    technical_snapshot_date,
                    run_id,
                    ready_documents,
                )
                matched_rows.append((registry_row, matches))

            print(summarize(registry_rows, matched_rows))

            if args.command == "dry-run":
                return 0

            row_id_map = upsert_raw_rows(
                cursor,
                registry_rows,
                source_system,
                source_snapshot_label,
                technical_snapshot_date,
                str(workbook_path),
            )

            xwalk_rows: list[XwalkRow] = []
            for registry_row, matches in matched_rows:
                registry_raw_id = row_id_map[(registry_row.source_sheet, registry_row.source_row_num)]
                for match in matches:
                    match.registry_raw_id = registry_raw_id
                    xwalk_rows.append(match)

            refresh_auto_xwalk(cursor, xwalk_rows, source_system, technical_snapshot_date, run_id)
            connection.commit()

            print("")
            print(verify_loaded_state(cursor, source_system, technical_snapshot_date, run_id))

    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main(sys.argv))
    except KeyboardInterrupt:
        raise SystemExit(130)
